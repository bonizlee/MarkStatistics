#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
statistics the score GUI
file:MarkStatisticsW.py
author:BonizLee
"""
import json
import sys
from PyQt5.QtWidgets import QMainWindow, QApplication, QAction, QTextEdit, QMessageBox
from PyQt5.QtGui import QIcon
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os.path
import images_qr
from PyQt5.QtCore import QCoreApplication

class MarkStatisticsW(QMainWindow):  

    def __init__(self):
        super().__init__()
        self.initUI()
        self.initConfig()
         
    def initUI(self):
        self.setWindowTitle('自动分数统计')
        self.setGeometry(300, 300, 500, 400)
        self.setWindowIcon(QIcon(':/ms.ico'))
        self.setStatusBar('加载配置完成')
        
        self.calcAction = QAction(QIcon(':/Calc.png'), '点击计算', self)
        self.calcAction.setStatusTip('点击进行计算')
        self.calcAction.triggered.connect(self.calc)
        self.toolbar = self.addToolBar('点击计算')
        self.toolbar.addAction(self.calcAction)

        self.textMessage = QTextEdit()
        self.textMessage.setReadOnly(True)
        self.setCentralWidget(self.textMessage)
        self.show()
        
    # 设置状态栏
    def setStatusBar(self, message):
        self.statusBar().showMessage(message)
    

    def errorDialog(self, msg):
        reply = QMessageBox.warning(self, "错误", msg, QMessageBox.Ok)
        

    # 计算
    def calc(self):
        self.summary()
        self.setStatusBar('汇总统计')
        self.write_excel()
        self.setStatusBar('保存完成')

    #初始化配置
    def initConfig(self):   
        global PATH
        PATH = os.path.dirname(os.path.realpath(__file__))+os.path.sep
        try:
            fp = open(PATH+'config.json') 
        except IOError:
            self.errorDialog('打开配置文件错误')            
            QCoreApplication.quit()

        global COMMOM_DATA
        try:
            COMMOM_DATA = json.load(fp)
        except Exception:
            self.errorDialog('读取配置错误')            
            QCoreApplication.quit()
        
        self.textMessage.append('加载配置文件内容:')
        self.textMessage.append(str(COMMOM_DATA))
        global MAXNUMBER 
        MAXNUMBER = COMMOM_DATA['maxnumber']
        global PROJECT 
        PROJECT = COMMOM_DATA['project']
        global STUDENT_MARK
        STUDENT_MARK = [0 for i in range(MAXNUMBER)]
        global FILETYPE
        FILETYPE = '.'+COMMOM_DATA['filetype']

    #读取统计
    def summary(self):      
        for subject in COMMOM_DATA['subject']:        
            judges = subject['judges']
            subjectmark =  [ [ 0 for i in range(MAXNUMBER) ] for j in range(judges) ]
            for n in range(0,judges):            
                filename = PATH+PROJECT+subject['filename']+str(n+1)+FILETYPE
                try:
                    workbook = load_workbook(filename,data_only=True)
                    sheetnames =workbook.get_sheet_names()
                    sheet = workbook.get_sheet_by_name(sheetnames[0])
                except Exception:
                    self.errorDialog('读取'+filename+'错误')
                    QCoreApplication.quit()

                start_cell = sheet[subject['markcell']]
                start_col = column_index_from_string(start_cell.column)
                start_row = start_cell.row
                for i in range(0, MAXNUMBER):
                    markcell = sheet.cell(row=start_row, column=start_col+i)               
                    subjectmark[n][i] = markcell.value
            self.textMessage.append('评分表'+subject['filename']+'完成')
            self.textMessage.append(str(subjectmark))
            calc_type = subject['calculate']
            if calc_type == 1:
                self.average(subjectmark, judges)
            elif calc_type == 2:
                self.without_max_min_average(subjectmark, judges)
            elif calc_type == 3:
                self.without_abs_max_average(subjectmark, judges)

    #算术平均
    def average(self, subjectmark, judges):
        for i in range(0, MAXNUMBER):
            sum = 0
            for n in range(0, judges):
                sum += subjectmark[n][i]
            STUDENT_MARK[i] += sum / judges

    #去除最高和最低再求平均
    def without_max_min_average(self, subjectmark, judges):
        for i in range(0, MAXNUMBER):
            max = subjectmark[0][i]
            sum = subjectmark[0][i]
            min = subjectmark[0][i]
            for n in range(1, judges):
                value = subjectmark[n][i]
                sum += value
                if value > max:
                    max = value
                elif value < min:
                    min = value
            STUDENT_MARK[i] += (sum - min -max) / (judges -2)

    #去除偏差最大的再求平均
    def without_abs_max_average(self, subjectmark, judges):
        for i in range(0, MAXNUMBER):
            jm = [0 for k in range(judges)]
            for n in range(0, judges):
                jm[n] = subjectmark[n][i]
            average_mark = sum(jm) / len(jm)
            max_abs = abs(average_mark-jm[0])
            max_mark = jm[0]
            for j in range(1, judges):
                if max_abs < abs(average_mark-jm[j]) and max_mark > jm[j]:
                    max_abs = abs(average_mark-jm[j])
                    max_mark = jm[j]
            STUDENT_MARK[i] += (sum(jm) - max_mark)/(judges - 1)

    #写入excel文件
    def write_excel( self):
        filename = PATH+PROJECT + 'summary'+FILETYPE   
        workbook = Workbook()
        sheet = workbook.get_active_sheet()
        sheet.cell(row=1, column=1).value = '工位号'
        sheet.cell(row=1, column=2).value = '成绩'
        sheet.cell(row=1, column=3).value = '名次'
        first_mark_cell = sheet.cell(row=2, column=2)
        last_mark_cell = sheet.cell(row=len(STUDENT_MARK)+1, column=2)
        self.textMessage.append('成绩:')
        self.textMessage.append(str(STUDENT_MARK))

        for i in range(len(STUDENT_MARK)):
            sheet.cell(row=2+i, column=1).value = i+1
            sheet.cell(row=2+i, column=2).number_format = '0.0'
            sheet.cell(row=2+i, column=2).value = STUDENT_MARK[i]
            formual_str = '=rank('+sheet.cell(row=2+i, column=2).coordinate + ',' +\
            first_mark_cell.column + '$2:' +\
            last_mark_cell.column +'$'+str(last_mark_cell.row)+ ')'
            sheet.cell(row=2+i, column=3).value = formual_str
        try:
            workbook.save(filename)
        except Exception:
            self.errorDialog('写入Excel错误')        

if __name__ == "__main__":
    app = QApplication(sys.argv)
    msw = MarkStatisticsW()
    sys.exit(app.exec_())