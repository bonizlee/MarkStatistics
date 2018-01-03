# statistics the score
# MarkStatistics.py
# BonizLee
# -*- coding: utf-8 -*-

import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os.path

#初始化
def init():   
    global PATH
    PATH = os.path.dirname(os.path.realpath(__file__))+os.path.sep
    fp = open(PATH+'config.json',encoding='utf-8')    #先以utf-8形式读入json文件
    global COMMOM_DATA
    jsonstr=fp.read()
    COMMOM_DATA = json.loads(jsonstr)  #loads方法为将字符串转换为json对象
    print(COMMOM_DATA)
    global MAXNUMBER 
    MAXNUMBER = COMMOM_DATA['maxnumber']
    global PROJECT 
    PROJECT = COMMOM_DATA['project']
    print(PROJECT)
    global STUDENT_MARK
    STUDENT_MARK = [0 for i in range(MAXNUMBER)]
    global FILETYPE
    FILETYPE = '.'+COMMOM_DATA['filetype']

#读取统计
def summary():      
    for subject in COMMOM_DATA['subject']:        
        judges = subject['judges']
        subjectmark =  [ [ 0 for i in range(MAXNUMBER) ] for j in range(judges) ]
        for n in range(0,judges):            
            filename = PATH+PROJECT+subject['filename']+str(n+1)+FILETYPE
            workbook = load_workbook(filename,data_only=True)
            sheetnames =workbook.get_sheet_names()
            sheet = workbook.get_sheet_by_name(sheetnames[0])
            start_cell = sheet[subject['markcell']]
            start_col = column_index_from_string(start_cell.column)
            start_row = start_cell.row
            for i in range(0,MAXNUMBER):
                markcell = sheet.cell(row=start_row,column=start_col+i)                
                subjectmark[n][i] = markcell.value
        print('评分表'+subject['filename']+'完成')
        print(subjectmark)
        calc_type = subject['calculate']
        if calc_type == 1:
            average(subjectmark,judges)
        elif calc_type == 2:
            without_max_min_average(subjectmark,judges)
        elif calc_type == 3:
            without_abs_max_average(subjectmark,judges)     

#算术平均
def average( subjectmark,judges ):    
    for i in range(0,MAXNUMBER):
        sum = 0
        for n in range(0,judges):
            sum += subjectmark[n][i]
        STUDENT_MARK[i] += sum / judges

#去除最高和最低再求平均
def without_max_min_average(  subjectmark,judges ):
    for i in range(0,MAXNUMBER):
        max = subjectmark[0][i]
        sum = subjectmark[0][i]
        min = subjectmark[0][i]
        for n in range(1,judges):
            value = subjectmark[n][i]
            sum += value
            if value > max:
                max = value
            elif value < min:
                min = value       
        STUDENT_MARK[i] += (sum - min -max) / (judges -2)

#去除偏差最大的再求平均
def without_abs_max_average( subjectmark,judges ):    
    for i in range(0,MAXNUMBER):
        jm = [0 for k in range(judges)]
        for n in range(0,judges):
            jm[n] = subjectmark[n][i]
        average_mark = sum(jm) / len(jm)               
        max_abs = abs(average_mark-jm[0])
        max_mark = jm[0]
        for j in range(1,judges):
            if max_abs < abs(average_mark-jm[j]) and max_mark > jm[j]:
                max_abs = abs(average_mark-jm[j])
                max_mark = jm[j]
        STUDENT_MARK[i] += (sum(jm) - max_mark)/(judges - 1)
        
#写入excel文件
def write_excel():
    filename = PATH+PROJECT + 'summary'+FILETYPE   
    workbook = Workbook()
    sheet = workbook.get_active_sheet()
    sheet.cell(row=1,column=1).value = '工位号'
    sheet.cell(row=1,column=2).value = '成绩'
    sheet.cell(row=1,column=3).value = '名次'
    first_mark_cell = sheet.cell(row=2,column=2)
    last_mark_cell = sheet.cell(row=len(STUDENT_MARK)+1,column=2)
    for i in range(len(STUDENT_MARK)):
        sheet.cell(row=2+i,column=1).value = i+1
        sheet.cell(row=2+i,column=2).number_format = '0.0'
        sheet.cell(row=2+i,column=2).value = STUDENT_MARK[i]
        formual_str = '=rank('+sheet.cell(row=2+i,column=2).coordinate + ',' +\
        first_mark_cell.column + '$2:' +\
        last_mark_cell.column +'$'+str(last_mark_cell.row)+ ')'
        sheet.cell(row=2+i,column=3).value = formual_str
    workbook.save(filename)

if __name__ == "__main__":
    init()
    print('加载配置完成')
    summary()
    print('汇总完成完成')
    print(STUDENT_MARK)
    write_excel()
    print('保存完成')
    while True:
        q = input("输入Q退出程序：")
        if q == 'Q' or q == 'q':
            break 
